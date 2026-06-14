"""Report service layer."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any

from django.core.cache import cache
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook


@dataclass(frozen=True)
class CachedReport:
    key: str
    timeout: int = 60


def get_cached_report(key: str):
    return cache.get(key)


def set_cached_report(key: str, value: Any, timeout: int = 60):
    cache.set(key, value, timeout)
    return value


def cache_key(namespace: str, canteen_id: str, params: dict[str, Any]) -> str:
    normalized = json.dumps(params, sort_keys=True, default=str)
    return f"cms:report:{namespace}:{canteen_id}:{normalized}"


def csv_response(filename: str, rows: list[dict[str, Any]], fieldnames: list[str]) -> HttpResponse:
    excel_filename = filename.rsplit(".", 1)[0] + ".xlsx"
    return excel_response(excel_filename, rows, fieldnames)


def excel_response(filename: str, rows: list[dict[str, Any]], fieldnames: list[str]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Spending"

    # Write headers
    for col_num, header in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, row in enumerate(rows, 2):
        for col_num, field in enumerate(fieldnames, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(field, ""))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def envelope(results, count: int, **extra):
    payload = {
        "results": results,
        "count": count,
        "generatedAt": timezone.now().isoformat(),
    }
    payload.update(extra)
    return payload
